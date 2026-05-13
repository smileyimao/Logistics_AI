"""
打包后双击运行，程序藏在系统托盘（右下角）。
有新运单自动写入 Excel，并弹出通知。
右键托盘图标可打开 Excel 或退出。
"""
import os
import sys
import time
import threading
import requests
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from PIL import Image, ImageDraw
import pystray

SERVER_URL = "http://localhost:5001/get_new_data"   # 本地测试用，上线后改成真实域名
EXCEL_PATH = os.path.join(os.path.expanduser("~"), "Desktop", "运单数据.xlsx")
HEADERS = ["运单号", "发货人", "收货人", "收货地址", "重量", "件数", "货物品名", "备注", "录入时间"]


# ── Excel ───────────────────────────────────────────────

def get_or_create_wb():
    if os.path.exists(EXCEL_PATH):
        return openpyxl.load_workbook(EXCEL_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "运单记录"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(EXCEL_PATH)
    return wb


def write_row(data: dict) -> str:
    wb = get_or_create_wb()
    ws = wb.active
    ws.append([
        data.get("运单号", ""),
        data.get("发货人", ""),
        data.get("收货人", ""),
        data.get("收货地址", ""),
        data.get("重量", ""),
        data.get("件数", ""),
        data.get("货物品名", ""),
        data.get("备注", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    wb.save(EXCEL_PATH)
    return data.get("运单号") or "未知单号"


# ── 托盘图标 ─────────────────────────────────────────────

def make_icon(color=(34, 197, 94)):
    """画一个纯色圆形图标"""
    img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
    ImageDraw.Draw(img).ellipse([4, 4, 60, 60], fill=(*color, 255))
    return img


def open_excel(icon, _):
    if sys.platform == "win32":
        os.startfile(EXCEL_PATH)
    else:
        os.system(f"open '{EXCEL_PATH}'")


def quit_app(icon, _):
    icon.stop()
    sys.exit(0)


# ── 轮询线程 ─────────────────────────────────────────────

def polling_loop(icon):
    while True:
        try:
            items = requests.get(SERVER_URL, timeout=5).json()
            for item in items:
                no = write_row(item["data"])
                try:
                    icon.notify(f"已写入运单：{no}", "运单助手")
                except Exception:
                    pass
        except requests.exceptions.ConnectionError:
            pass
        except Exception:
            pass
        time.sleep(5)


# ── 入口 ─────────────────────────────────────────────────

menu = pystray.Menu(
    pystray.MenuItem("打开 Excel", open_excel),
    pystray.MenuItem("退出", quit_app),
)

icon = pystray.Icon(
    name="运单助手",
    icon=make_icon(),
    title="运单助手 — 运行中",
    menu=menu,
)

threading.Thread(target=polling_loop, args=(icon,), daemon=True).start()

icon.run()
